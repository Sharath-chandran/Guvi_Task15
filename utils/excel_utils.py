import openpyxl
from datetime import datetime

class ExcelUtil:
    # Initializes the class with the path to the Excel file.
    # The path to the Excel file you want to read.
    def __init__(self, file_path):
        # Stores the file_path in an instance variable self.file_path so it can be used later within the class.
        self.file_path = file_path
        # Uses the openpyxl library to load the Excel workbook from the given file path.
        # self.workbook now holds the entire Excel file in memory.
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active # Selects the active sheet in the workbook (usually the first sheet that opens by default).

    # Method named get_data that will extract data from the Excel sheet.
    def get_data(self):
        data = [] # Initializes an empty list called data to store the extracted rows.
        # Loops through each row in the sheet starting from the second row (min_row=2) to skip the header.
        for row in self.sheet.iter_rows(min_row=2, values_only=False): # iter_rows returns each row as a tuple of cell objects.
            test_id = row[0].value # Extracts the value from the first cell in the row (column A)
            username = row[1].value # Extracts the value from the second cell in the row (column B)
            password = row[2].value # Extracts the value from the third cell in the row (column C)
            # Appends a tuple of the three values (test_id, username, password) to the data list.
            data.append((test_id, username, password))
        return data # Returns the complete list of tuples containing the extracted data.

    def write_result(self, row, result, tester_name="Sharath"):
        now = datetime.now() # Gets the current date and time
        self.sheet.cell(row=row, column=4).value = now.date().isoformat() #  formats the date as a string like "yyyy-mm-dd".
        self.sheet.cell(row=row, column=5).value = now.time().strftime("%H:%M:%S") # formats the time as "HH:MM:SS"
        self.sheet.cell(row=row, column=6).value = tester_name
        self.sheet.cell(row=row, column=7).value = result
        self.workbook.save(self.file_path)